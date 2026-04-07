"""Spreadsheet tools — read, write, diff, and inspect Excel (.xlsx) and CSV files.

Enables the AI-human handoff loop: agent produces a spreadsheet → user edits in
Excel → agent reads it back, diffs changes, and continues.  Preserves formulas,
formatting metadata, and multi-sheet structure across round-trips.

Dependencies: openpyxl (XLSX), csv (stdlib).
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from isaac.core.types import PermissionLevel, ToolDef


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F811
        return openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel support. Install it: pip install openpyxl"
        )


def _file_fingerprint(path: Path) -> dict[str, Any]:
    """Return size, mtime, and content hash for change detection."""
    stat = path.stat()
    content_hash = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
    return {
        "size_bytes": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "sha256_prefix": content_hash,
    }


def _cell_to_str(cell) -> str:
    """Convert an openpyxl cell to a display string, preserving type info."""
    if cell.value is None:
        return ""
    return str(cell.value)


def _cell_detail(cell) -> dict[str, Any]:
    """Extract rich cell metadata for round-trip fidelity."""
    detail: dict[str, Any] = {"value": cell.value}
    if cell.data_type == "f":
        detail["formula"] = cell.value
    if cell.number_format and cell.number_format != "General":
        detail["number_format"] = cell.number_format
    if cell.comment:
        detail["comment"] = cell.comment.text
    return detail


def _read_xlsx(path: Path, sheet: str | None, max_rows: int, max_cols: int,
               include_formulas: bool) -> dict[str, Any]:
    """Read an XLSX file into a structured representation."""
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(str(path), data_only=not include_formulas)

    sheet_names = wb.sheetnames
    target_sheets = [sheet] if sheet and sheet in sheet_names else sheet_names

    result: dict[str, Any] = {
        "path": str(path),
        "format": "xlsx",
        "sheet_names": sheet_names,
        "fingerprint": _file_fingerprint(path),
        "sheets": {},
    }

    for sname in target_sheets:
        ws = wb[sname]
        rows_data: list[list[Any]] = []
        header: list[str] = []
        formulas_found: list[dict[str, Any]] = []

        effective_cols = min(max_cols, ws.max_column or max_cols)
        effective_rows = min(max_rows, ws.max_row or max_rows)
        for r_idx, row in enumerate(ws.iter_rows(max_row=effective_rows, max_col=effective_cols), 1):
            row_vals: list[Any] = []
            for c_idx, cell in enumerate(row, 1):
                if include_formulas and cell.data_type == "f":
                    formulas_found.append({
                        "cell": cell.coordinate,
                        "formula": str(cell.value),
                    })
                row_vals.append(_cell_to_str(cell))
            if r_idx == 1:
                header = [str(v) for v in row_vals]
            rows_data.append(row_vals)

        sheet_info: dict[str, Any] = {
            "dimensions": ws.dimensions or "",
            "row_count": ws.max_row or 0,
            "col_count": ws.max_column or 0,
            "header": header,
            "data": rows_data,
            "rows_returned": len(rows_data),
            "truncated": (ws.max_row or 0) > max_rows or (ws.max_column or 0) > max_cols,
        }
        if formulas_found:
            sheet_info["formulas"] = formulas_found

        # Capture merged cells
        if ws.merged_cells.ranges:
            sheet_info["merged_cells"] = [str(r) for r in ws.merged_cells.ranges]

        result["sheets"][sname] = sheet_info

    wb.close()
    return result


def _read_csv(path: Path, max_rows: int, max_cols: int) -> dict[str, Any]:
    """Read a CSV file into a structured representation."""
    rows_data: list[list[str]] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows_data.append(row[:max_cols])

    header = rows_data[0] if rows_data else []
    total_rows = sum(1 for _ in open(path, encoding="utf-8-sig"))

    return {
        "path": str(path),
        "format": "csv",
        "fingerprint": _file_fingerprint(path),
        "sheets": {
            "Sheet1": {
                "header": header,
                "data": rows_data,
                "row_count": total_rows,
                "col_count": len(header),
                "rows_returned": len(rows_data),
                "truncated": total_rows > max_rows,
            },
        },
    }


def build_spreadsheet_tools(
    cwd: str | None = None,
) -> dict[str, tuple[ToolDef, Any]]:
    """Build spreadsheet tools for Excel/CSV reading, writing, diffing, and inspection."""
    working_dir = cwd or os.getcwd()
    registry: dict[str, tuple[ToolDef, Any]] = {}

    def _resolve(path: str) -> Path:
        p = Path(path)
        return p if p.is_absolute() else Path(working_dir) / p

    # ── spreadsheet_read ──────────────────────────────────────────────

    async def spreadsheet_read(
        path: str,
        sheet: str = "",
        max_rows: int = 500,
        max_cols: int = 50,
        include_formulas: bool = True,
    ) -> dict[str, Any]:
        """Read an Excel or CSV file into a structured JSON representation.

        Returns headers, rows, formulas, merged cells, and a fingerprint for
        change detection on re-read.
        """
        target = _resolve(path)
        if not target.exists():
            return {"error": f"File not found: {target}"}

        suffix = target.suffix.lower()
        try:
            if suffix in (".xlsx", ".xlsm", ".xltx"):
                return _read_xlsx(
                    target,
                    sheet=sheet or None,
                    max_rows=max_rows,
                    max_cols=max_cols,
                    include_formulas=include_formulas,
                )
            elif suffix == ".csv":
                return _read_csv(target, max_rows=max_rows, max_cols=max_cols)
            else:
                return {"error": f"Unsupported format: {suffix}. Use .xlsx, .xlsm, .xltx, or .csv"}
        except Exception as e:
            return {"error": f"Failed to read {target.name}: {e}"}

    registry["spreadsheet_read"] = (
        ToolDef(
            name="spreadsheet_read",
            description=(
                "Read an Excel (.xlsx) or CSV file. Returns structured data: headers, rows, "
                "formulas, merged cells, sheet names, and a fingerprint for tracking changes. "
                "Use this instead of file_read for spreadsheets — it understands the structure."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to .xlsx or .csv file"},
                    "sheet": {"type": "string", "description": "Sheet name (XLSX only, empty = all sheets)", "default": ""},
                    "max_rows": {"type": "integer", "description": "Max rows to return", "default": 500},
                    "max_cols": {"type": "integer", "description": "Max columns to return", "default": 50},
                    "include_formulas": {"type": "boolean", "description": "Show formulas instead of computed values", "default": True},
                },
                "required": ["path"],
            },
            permission=PermissionLevel.AUTO,
            is_read_only=True,
        ),
        spreadsheet_read,
    )

    # ── spreadsheet_write ─────────────────────────────────────────────

    async def spreadsheet_write(
        path: str,
        data: list[list[Any]],
        sheet: str = "Sheet1",
        format: str = "",
        header_style: bool = True,
    ) -> dict[str, Any]:
        """Write tabular data to an Excel or CSV file.

        data: list of rows, where the first row is treated as the header.
        format: 'xlsx' or 'csv'. Auto-detected from path extension if empty.
        """
        target = _resolve(path)
        suffix = format.lower() if format else target.suffix.lower().lstrip(".")
        if suffix not in ("xlsx", "csv"):
            return {"error": f"Unsupported format: {suffix}. Use xlsx or csv."}

        target.parent.mkdir(parents=True, exist_ok=True)

        try:
            if suffix == "csv":
                with open(target, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                return {
                    "written": str(target),
                    "format": "csv",
                    "rows": len(data),
                    "fingerprint": _file_fingerprint(target),
                }
            else:
                openpyxl = _require_openpyxl()
                from openpyxl.styles import Font

                # If the file already exists, open it to preserve other sheets
                if target.exists():
                    wb = openpyxl.load_workbook(str(target))
                    if sheet in wb.sheetnames:
                        del wb[sheet]
                    ws = wb.create_sheet(sheet)
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = sheet

                for r_idx, row in enumerate(data, 1):
                    for c_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        # Bold headers
                        if r_idx == 1 and header_style:
                            cell.font = Font(bold=True)

                wb.save(str(target))
                wb.close()
                return {
                    "written": str(target),
                    "format": "xlsx",
                    "sheet": sheet,
                    "rows": len(data),
                    "fingerprint": _file_fingerprint(target),
                }
        except Exception as e:
            return {"error": f"Failed to write {target.name}: {e}"}

    registry["spreadsheet_write"] = (
        ToolDef(
            name="spreadsheet_write",
            description=(
                "Write tabular data to an Excel (.xlsx) or CSV file. First row is treated as "
                "the header. For XLSX, preserves existing sheets — only the target sheet is "
                "overwritten. Returns a fingerprint for handoff tracking."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Output file path (.xlsx or .csv)"},
                    "data": {
                        "type": "array",
                        "items": {"type": "array", "items": {}},
                        "description": "Row data — list of lists. First row = header.",
                    },
                    "sheet": {"type": "string", "description": "Sheet name (XLSX only)", "default": "Sheet1"},
                    "format": {"type": "string", "description": "Force format: 'xlsx' or 'csv'. Auto-detected from extension if empty.", "default": ""},
                    "header_style": {"type": "boolean", "description": "Bold the header row in XLSX", "default": True},
                },
                "required": ["path", "data"],
            },
            permission=PermissionLevel.AUTO,
        ),
        spreadsheet_write,
    )

    # ── spreadsheet_edit ──────────────────────────────────────────────

    async def spreadsheet_edit(
        path: str,
        edits: list[dict[str, Any]],
        sheet: str = "",
    ) -> dict[str, Any]:
        """Apply surgical edits to specific cells in an existing XLSX file.

        edits: list of {"cell": "B3", "value": 42} or {"cell": "C5", "formula": "=SUM(C2:C4)"}
        Preserves all existing data, formatting, and formulas outside the edited cells.
        """
        target = _resolve(path)
        if not target.exists():
            return {"error": f"File not found: {target}"}

        suffix = target.suffix.lower()
        if suffix not in (".xlsx", ".xlsm", ".xltx"):
            return {"error": "spreadsheet_edit only supports XLSX files. Use spreadsheet_write for CSV."}

        openpyxl = _require_openpyxl()
        try:
            wb = openpyxl.load_workbook(str(target))
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

            applied: list[str] = []
            for edit in edits:
                cell_ref = edit.get("cell", "")
                if not cell_ref:
                    continue
                cell = ws[cell_ref]
                if "formula" in edit:
                    cell.value = edit["formula"]
                    applied.append(f"{cell_ref}={edit['formula']}")
                elif "value" in edit:
                    cell.value = edit["value"]
                    applied.append(f"{cell_ref}={edit['value']}")

            wb.save(str(target))
            wb.close()
            return {
                "path": str(target),
                "edits_applied": len(applied),
                "details": applied,
                "fingerprint": _file_fingerprint(target),
            }
        except Exception as e:
            return {"error": f"Failed to edit {target.name}: {e}"}

    registry["spreadsheet_edit"] = (
        ToolDef(
            name="spreadsheet_edit",
            description=(
                "Apply surgical edits to specific cells in an XLSX file without touching "
                "anything else. Preserves formatting, formulas, and other sheets. "
                "Use this for targeted updates instead of rewriting the whole file."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to existing .xlsx file"},
                    "edits": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "cell": {"type": "string", "description": "Cell reference (e.g. 'B3', 'AA12')"},
                                "value": {"description": "New value for the cell"},
                                "formula": {"type": "string", "description": "Formula (e.g. '=SUM(A1:A10)')"},
                            },
                            "required": ["cell"],
                        },
                        "description": "List of cell edits. Each must have 'cell' and either 'value' or 'formula'.",
                    },
                    "sheet": {"type": "string", "description": "Target sheet (empty = active sheet)", "default": ""},
                },
                "required": ["path", "edits"],
            },
            permission=PermissionLevel.AUTO,
        ),
        spreadsheet_edit,
    )

    # ── spreadsheet_diff ──────────────────────────────────────────────

    async def spreadsheet_diff(
        path_a: str,
        path_b: str,
        sheet: str = "",
        max_diffs: int = 200,
    ) -> dict[str, Any]:
        """Compare two spreadsheet files cell-by-cell.

        Returns a list of changed, added, and removed cells — the core primitive
        for the human-AI handoff loop.  After a user edits a spreadsheet in Excel,
        diff it against the agent's version to see exactly what changed.
        """
        target_a = _resolve(path_a)
        target_b = _resolve(path_b)
        for t in (target_a, target_b):
            if not t.exists():
                return {"error": f"File not found: {t}"}

        try:
            # Read both files
            read_a = await spreadsheet_read(
                str(target_a), sheet=sheet, max_rows=2000, max_cols=100, include_formulas=True,
            )
            read_b = await spreadsheet_read(
                str(target_b), sheet=sheet, max_rows=2000, max_cols=100, include_formulas=True,
            )

            if "error" in read_a:
                return read_a
            if "error" in read_b:
                return read_b

            diffs: list[dict[str, Any]] = []
            sheets_a = read_a.get("sheets", {})
            sheets_b = read_b.get("sheets", {})

            all_sheets = set(sheets_a.keys()) | set(sheets_b.keys())
            for sname in sorted(all_sheets):
                if sname not in sheets_a:
                    diffs.append({"sheet": sname, "change": "sheet_added"})
                    continue
                if sname not in sheets_b:
                    diffs.append({"sheet": sname, "change": "sheet_removed"})
                    continue

                data_a = sheets_a[sname].get("data", [])
                data_b = sheets_b[sname].get("data", [])
                max_r = max(len(data_a), len(data_b))
                max_c = max(
                    (max(len(r) for r in data_a) if data_a else 0),
                    (max(len(r) for r in data_b) if data_b else 0),
                )

                for r in range(max_r):
                    for c in range(max_c):
                        val_a = data_a[r][c] if r < len(data_a) and c < len(data_a[r]) else ""
                        val_b = data_b[r][c] if r < len(data_b) and c < len(data_b[r]) else ""
                        if val_a != val_b:
                            # Convert to Excel-style cell ref
                            col_letter = ""
                            col_num = c
                            while col_num >= 0:
                                col_letter = chr(65 + col_num % 26) + col_letter
                                col_num = col_num // 26 - 1
                            cell_ref = f"{col_letter}{r + 1}"

                            diffs.append({
                                "sheet": sname,
                                "cell": cell_ref,
                                "old": val_a,
                                "new": val_b,
                            })
                            if len(diffs) >= max_diffs:
                                break
                    if len(diffs) >= max_diffs:
                        break

            return {
                "path_a": str(target_a),
                "path_b": str(target_b),
                "fingerprint_a": read_a.get("fingerprint"),
                "fingerprint_b": read_b.get("fingerprint"),
                "total_diffs": len(diffs),
                "truncated": len(diffs) >= max_diffs,
                "diffs": diffs,
            }
        except Exception as e:
            return {"error": f"Diff failed: {e}"}

    registry["spreadsheet_diff"] = (
        ToolDef(
            name="spreadsheet_diff",
            description=(
                "Compare two spreadsheet files cell-by-cell. Returns exactly which cells "
                "changed, were added, or removed. Essential for the human-AI handoff: after a "
                "user edits a file in Excel, diff it against the original to see what changed."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path_a": {"type": "string", "description": "Path to the original/baseline file"},
                    "path_b": {"type": "string", "description": "Path to the modified file"},
                    "sheet": {"type": "string", "description": "Compare a specific sheet (empty = all)", "default": ""},
                    "max_diffs": {"type": "integer", "description": "Max cell diffs to return", "default": 200},
                },
                "required": ["path_a", "path_b"],
            },
            permission=PermissionLevel.AUTO,
            is_read_only=True,
        ),
        spreadsheet_diff,
    )

    # ── spreadsheet_inspect ───────────────────────────────────────────

    async def spreadsheet_inspect(path: str) -> dict[str, Any]:
        """Get metadata about a spreadsheet without reading all its data.

        Returns: sheet names, dimensions, formula count, named ranges,
        file fingerprint.  Useful for understanding a file before deciding
        what to read.
        """
        target = _resolve(path)
        if not target.exists():
            return {"error": f"File not found: {target}"}

        suffix = target.suffix.lower()
        if suffix == ".csv":
            total_rows = sum(1 for _ in open(target, encoding="utf-8-sig"))
            with open(target, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                first_row = next(reader, [])
            return {
                "path": str(target),
                "format": "csv",
                "fingerprint": _file_fingerprint(target),
                "row_count": total_rows,
                "col_count": len(first_row),
                "header": first_row,
            }

        if suffix not in (".xlsx", ".xlsm", ".xltx"):
            return {"error": f"Unsupported format: {suffix}"}

        openpyxl = _require_openpyxl()
        try:
            wb = openpyxl.load_workbook(str(target), read_only=True)
            sheets_info: dict[str, Any] = {}
            for sname in wb.sheetnames:
                ws = wb[sname]
                sheets_info[sname] = {
                    "row_count": ws.max_row or 0,
                    "col_count": ws.max_column or 0,
                }
            result = {
                "path": str(target),
                "format": "xlsx",
                "fingerprint": _file_fingerprint(target),
                "sheet_names": wb.sheetnames,
                "sheets": sheets_info,
            }
            # Named ranges
            if wb.defined_names:
                result["named_ranges"] = [
                    {"name": dn.name, "value": str(dn.value)}
                    for dn in wb.defined_names.definedName
                ]
            wb.close()
            return result
        except Exception as e:
            return {"error": f"Failed to inspect {target.name}: {e}"}

    registry["spreadsheet_inspect"] = (
        ToolDef(
            name="spreadsheet_inspect",
            description=(
                "Get metadata about a spreadsheet without reading the data. Returns sheet "
                "names, dimensions, row/column counts, named ranges, and fingerprint. "
                "Use this first to understand a large file before deciding what to read."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to .xlsx or .csv file"},
                },
                "required": ["path"],
            },
            permission=PermissionLevel.AUTO,
            is_read_only=True,
        ),
        spreadsheet_inspect,
    )

    # ── spreadsheet_convert ───────────────────────────────────────────

    async def spreadsheet_convert(
        path: str,
        output_format: str,
        output_path: str = "",
        sheet: str = "",
    ) -> dict[str, Any]:
        """Convert between XLSX and CSV formats.

        XLSX → CSV: exports the specified sheet (or active sheet).
        CSV → XLSX: imports into a single sheet.
        """
        target = _resolve(path)
        if not target.exists():
            return {"error": f"File not found: {target}"}

        out_fmt = output_format.lower().lstrip(".")
        if out_fmt not in ("xlsx", "csv"):
            return {"error": f"Unsupported output format: {out_fmt}. Use xlsx or csv."}

        in_suffix = target.suffix.lower()
        if in_suffix == f".{out_fmt}":
            return {"error": f"Input and output formats are the same: {out_fmt}"}

        out_target = _resolve(output_path) if output_path else target.with_suffix(f".{out_fmt}")
        out_target.parent.mkdir(parents=True, exist_ok=True)

        try:
            # Read the source
            read_result = await spreadsheet_read(
                str(target), sheet=sheet, max_rows=100_000, max_cols=500, include_formulas=False,
            )
            if "error" in read_result:
                return read_result

            sheets = read_result.get("sheets", {})
            # Pick the sheet to convert
            if sheet and sheet in sheets:
                data = sheets[sheet]["data"]
            else:
                first_sheet = next(iter(sheets))
                data = sheets[first_sheet]["data"]

            # Write in the target format
            write_result = await spreadsheet_write(
                str(out_target), data=data, format=out_fmt,
            )
            if "error" in write_result:
                return write_result

            return {
                "source": str(target),
                "output": str(out_target),
                "source_format": in_suffix.lstrip("."),
                "output_format": out_fmt,
                "rows": len(data),
                "fingerprint": _file_fingerprint(out_target),
                "note": "Formulas converted to static values" if in_suffix in (".xlsx", ".xlsm") and out_fmt == "csv" else "",
            }
        except Exception as e:
            return {"error": f"Conversion failed: {e}"}

    registry["spreadsheet_convert"] = (
        ToolDef(
            name="spreadsheet_convert",
            description=(
                "Convert between XLSX and CSV formats. XLSX→CSV flattens formulas to values "
                "and exports one sheet. CSV→XLSX wraps in a proper workbook. "
                "Note: XLSX and CSV are NOT interchangeable — XLSX preserves formulas, "
                "formatting, multiple sheets, and types. CSV is plain text."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Source file path"},
                    "output_format": {"type": "string", "description": "'xlsx' or 'csv'"},
                    "output_path": {"type": "string", "description": "Output path (default: same name, new extension)", "default": ""},
                    "sheet": {"type": "string", "description": "Sheet to export (XLSX→CSV only)", "default": ""},
                },
                "required": ["path", "output_format"],
            },
            permission=PermissionLevel.AUTO,
        ),
        spreadsheet_convert,
    )

    return registry
